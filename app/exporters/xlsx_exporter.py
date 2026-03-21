"""Otomoto Parser – XLSX exporter using openpyxl."""
from __future__ import annotations

from datetime import date, datetime
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.storage.models import Listing

logger = logging.getLogger(__name__)

_COLUMNS = [
    ("otomoto_id", "ID"),
    ("title", "Tytuł"),
    ("price_pln", "Cena (PLN)"),
    ("year", "Rok"),
    ("mileage_km", "Przebieg (km)"),
    ("fuel_type", "Paliwo"),
    ("transmission", "Skrzynia"),
    ("engine_cc", "Pojemność (cc)"),
    ("power_hp", "Moc (KM)"),
    ("body_type", "Nadwozie"),
    ("colour", "Kolor"),
    ("condition", "Stan"),
    ("seller_type", "Sprzedający"),
    ("location", "Lokalizacja"),
    ("origin_country", "Kraj pochodzenia"),
    ("published_at", "Published at"),
    ("source_category", "Source category"),
    ("raw_attributes_json", "Raw attributes"),
    ("photos_json", "Photos JSON"),
    ("url", "Link"),
    ("scraped_at", "Data scrape"),
    ("updated_at", "Aktualizacja"),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
CellValue = str | int | float | bool | date | datetime


class XlsxExporter:
    """Exports Listing records to a formatted XLSX workbook."""

    def export(self, listings: list[Listing], path: Path) -> None:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Workbook has no active worksheet")
        ws.title = "Listings"

        # Header row
        for col_idx, (_, label) in enumerate(_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # Data rows
        for row_idx, listing in enumerate(listings, 2):
            fill = _ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, (field, _) in enumerate(_COLUMNS, 1):
                cell_value = self._to_cell_value(getattr(listing, field, ""))
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if fill:
                    cell.fill = fill
                if field == "url" and cell_value:
                    cell.hyperlink = str(cell_value)
                    cell.font = Font(color="0563C1", underline="single")

        # Auto-fit column widths
        for col_idx, (_, label) in enumerate(_COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(label)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        wb.save(path)
        logger.info(f"XLSX exported: {path} ({len(listings)} rows)")

    @staticmethod
    def _to_cell_value(value: object) -> CellValue | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, (str, int, float, bool)):
            return value
        return str(value)
